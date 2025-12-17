"""
Tests for ModelStorage

Comprehensive unit tests for model storage functionality including:
- save_checkpoint creates .pt and .json files
- load_checkpoint returns correct structure
- list_models returns all saved models
- delete_model removes files
- export_to_excel creates valid Excel with all sheets
"""

import pytest
import torch
import torch.nn as nn
import os
import json
import tempfile
import shutil
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import Dict, Any, List
from openpyxl import load_workbook

from app.services.model_storage import ModelStorage
from app.services.nn_scorecard import ScorecardNN


# ============================================================================
# MOCK METADATA STRUCTURE
# ============================================================================

@dataclass
class MockMetadata:
    """Mock metadata structure matching what ModelStorage expects."""
    model_id: str
    model_name: str
    created_at: str
    created_by: str
    segment: str
    architecture: Dict[str, Any]
    regularization: Dict[str, Any]
    loss_function: Dict[str, Any]
    optimizer: Dict[str, Any]
    training: Dict[str, Any]
    final_metrics: Dict[str, Any]
    training_history: Dict[str, Any]
    scorecard_output: Dict[str, Any]


def create_mock_metadata(model_id: str = "test-model-123", segment: str = "CONSUMER") -> MockMetadata:
    """Create a mock metadata object for testing."""
    return MockMetadata(
        model_id=model_id,
        model_name="Test Model",
        created_at="2024-01-15T10:30:00Z",
        created_by="test_user",
        segment=segment,
        architecture={
            "model_type": "neural_network",
            "input_dim": 10,
            "hidden_layers": [32, 16],
            "activation_function": "relu",
            "use_batch_normalization": True,
            "total_parameters": 1000
        },
        regularization={
            "dropout_rate": 0.2,
            "l1_lambda": 0.0,
            "l2_lambda": 0.01,
            "gradient_clip_norm": 1.0
        },
        loss_function={
            "loss_type": "combined",
            "loss_alpha": 0.3,
            "auc_gamma": 2.0
        },
        optimizer={
            "optimizer_type": "adam",
            "learning_rate": 0.001,
            "weight_decay": 0.0
        },
        training={
            "batch_size": 256,
            "epochs": 100,
            "early_stopping": {
                "enabled": True,
                "patience": 15,
                "min_delta": 0.001,
                "monitor": "test_ar"
            }
        },
        final_metrics={
            "discrimination": {
                "auc_roc": 0.85,
                "gini_ar": 0.70,
                "ks_statistic": 0.50
            },
            "calibration": {
                "brier_score": 0.12
            },
            "classification": {
                "accuracy": 0.80,
                "precision": 0.75,
                "recall": 0.70,
                "f1_score": 0.72
            }
        },
        training_history={
            "epochs": [
                {
                    "epoch": 1,
                    "train_loss": 0.6,
                    "test_loss": 0.65,
                    "train_auc": 0.70,
                    "test_auc": 0.68,
                    "train_ar": 0.40,
                    "test_ar": 0.36,
                    "train_ks": 0.35,
                    "test_ks": 0.32,
                    "learning_rate": 0.001
                },
                {
                    "epoch": 2,
                    "train_loss": 0.55,
                    "test_loss": 0.60,
                    "train_auc": 0.75,
                    "test_auc": 0.72,
                    "train_ar": 0.50,
                    "test_ar": 0.44,
                    "train_ks": 0.40,
                    "test_ks": 0.38,
                    "learning_rate": 0.001
                }
            ],
            "best_epoch": 2,
            "total_duration_seconds": 120.5
        },
        scorecard_output={
            "base_points": 50,
            "features": [
                {
                    "feature_name": "credit_score",
                    "model_weight": 0.45,
                    "bins": [
                        {
                            "bin_label": "Very Poor (<500)",
                            "woe_value": -0.823,
                            "points": 10
                        },
                        {
                            "bin_label": "Poor (500-580)",
                            "woe_value": -0.523,
                            "points": 25
                        },
                        {
                            "bin_label": "Good (580-650)",
                            "woe_value": 0.123,
                            "points": 50
                        },
                        {
                            "bin_label": "Excellent (>650)",
                            "woe_value": 0.623,
                            "points": 75
                        }
                    ]
                },
                {
                    "feature_name": "debt_to_income",
                    "model_weight": 0.35,
                    "bins": [
                        {
                            "bin_label": "High (>0.5)",
                            "woe_value": -0.456,
                            "points": 20
                        },
                        {
                            "bin_label": "Low (<0.5)",
                            "woe_value": 0.234,
                            "points": 60
                        }
                    ]
                }
            ]
        }
    )


# ============================================================================
# FIXTURES
# ============================================================================

@pytest.fixture
def temp_dir():
    """Create a temporary directory for testing."""
    temp_path = tempfile.mkdtemp()
    yield temp_path
    shutil.rmtree(temp_path)


@pytest.fixture
def storage(temp_dir):
    """Create a ModelStorage instance with temporary directory."""
    return ModelStorage(base_path=temp_dir)


@pytest.fixture
def sample_model():
    """Create a sample PyTorch model for testing."""
    return ScorecardNN(
        input_dim=10,
        hidden_layers=[32, 16],
        activation='relu',
        dropout_rate=0.2,
        use_batch_norm=True
    )


@pytest.fixture
def sample_metadata():
    """Create sample metadata for testing."""
    return create_mock_metadata()


# ============================================================================
# TEST 1: save_checkpoint creates .pt and .json files
# ============================================================================

def test_save_checkpoint_creates_pt_and_json_files(storage, sample_model):
    """Test that save_checkpoint creates both .pt and .json files."""
    model_id = "test-model-001"
    metadata = create_mock_metadata(model_id=model_id)
    
    # Save checkpoint
    checkpoint_path = storage.save_checkpoint(
        model=sample_model,
        metadata=metadata,
        model_id=model_id
    )
    
    # Check .pt file exists
    pt_path = os.path.join(storage.base_path, f"{model_id}.pt")
    assert os.path.exists(pt_path), f".pt file not found at {pt_path}"
    
    # Check .json file exists
    json_path = os.path.join(storage.base_path, f"{model_id}_metadata.json")
    assert os.path.exists(json_path), f".json file not found at {json_path}"
    
    # Verify checkpoint path matches expected
    assert checkpoint_path == pt_path
    
    # Verify .pt file contains valid checkpoint
    checkpoint = torch.load(pt_path, map_location='cpu', weights_only=False)
    assert 'model_state_dict' in checkpoint
    assert 'metadata' in checkpoint
    assert checkpoint['model_id'] == model_id
    assert checkpoint['segment'] == metadata.segment
    
    # Verify .json file contains valid JSON
    with open(json_path, 'r') as f:
        json_data = json.load(f)
    assert json_data['model_id'] == model_id
    assert json_data['segment'] == metadata.segment


# ============================================================================
# TEST 2: load_checkpoint returns correct structure
# ============================================================================

def test_load_checkpoint_returns_correct_structure(storage, sample_model):
    """Test that load_checkpoint returns the correct structure."""
    model_id = "test-model-002"
    metadata = create_mock_metadata(model_id=model_id)
    
    # Save checkpoint first
    storage.save_checkpoint(
        model=sample_model,
        metadata=metadata,
        model_id=model_id
    )
    
    # Load checkpoint
    result = storage.load_checkpoint(model_id)
    
    # Verify structure
    assert isinstance(result, dict), "Result should be a dictionary"
    assert 'model' in result, "Result should contain 'model' key"
    assert 'model_state_dict' in result, "Result should contain 'model_state_dict' key"
    assert 'metadata' in result, "Result should contain 'metadata' key"
    assert 'model_id' in result, "Result should contain 'model_id' key"
    assert 'segment' in result, "Result should contain 'segment' key"
    
    # Verify values
    assert result['model_id'] == model_id
    assert result['segment'] == metadata.segment
    assert result['metadata']['model_id'] == model_id
    assert result['metadata']['segment'] == metadata.segment
    
    # Verify model_state_dict is a dict
    assert isinstance(result['model_state_dict'], dict)
    
    # Test loading with model_class
    result_with_model = storage.load_checkpoint(
        model_id=model_id,
        model_class=ScorecardNN
    )
    
    # Verify model is instantiated
    assert result_with_model['model'] is not None
    assert isinstance(result_with_model['model'], nn.Module)


def test_load_checkpoint_raises_file_not_found(storage):
    """Test that load_checkpoint raises FileNotFoundError for non-existent model."""
    with pytest.raises(FileNotFoundError, match="Model not found"):
        storage.load_checkpoint("non-existent-model")


# ============================================================================
# TEST 3: list_models returns all saved models
# ============================================================================

def test_list_models_returns_all_saved_models(storage, sample_model):
    """Test that list_models returns all saved models."""
    # Create multiple models with different segments
    models = [
        ("model-001", "CONSUMER"),
        ("model-002", "SME"),
        ("model-003", "CONSUMER"),
    ]
    
    # Save all models
    for model_id, segment in models:
        metadata = create_mock_metadata(model_id=model_id, segment=segment)
        storage.save_checkpoint(
            model=sample_model,
            metadata=metadata,
            model_id=model_id
        )
    
    # List all models
    all_models = storage.list_models()
    
    # Verify we got all models
    assert len(all_models) == 3, f"Expected 3 models, got {len(all_models)}"
    
    # Verify structure of each model
    for model_info in all_models:
        assert 'model_id' in model_info
        assert 'segment' in model_info
        assert 'created_at' in model_info
        assert 'auc_roc' in model_info or model_info['auc_roc'] is None
        assert 'gini_ar' in model_info or model_info['gini_ar'] is None
    
    # Verify model IDs are present
    model_ids = [m['model_id'] for m in all_models]
    assert "model-001" in model_ids
    assert "model-002" in model_ids
    assert "model-003" in model_ids
    
    # Test filtering by segment
    consumer_models = storage.list_models(segment="CONSUMER")
    assert len(consumer_models) == 2, f"Expected 2 CONSUMER models, got {len(consumer_models)}"
    assert all(m['segment'] == "CONSUMER" for m in consumer_models)
    
    # Test filtering by non-existent segment
    empty_models = storage.list_models(segment="NONEXISTENT")
    assert len(empty_models) == 0


def test_list_models_returns_empty_list_when_no_models(storage):
    """Test that list_models returns empty list when no models exist."""
    models = storage.list_models()
    assert isinstance(models, list)
    assert len(models) == 0


# ============================================================================
# TEST 4: delete_model removes files
# ============================================================================

def test_delete_model_removes_files(storage, sample_model, sample_metadata):
    """Test that delete_model removes both .pt and .json files."""
    model_id = "test-model-004"
    
    # Save checkpoint first
    storage.save_checkpoint(
        model=sample_model,
        metadata=sample_metadata,
        model_id=model_id
    )
    
    # Verify files exist
    pt_path = os.path.join(storage.base_path, f"{model_id}.pt")
    json_path = os.path.join(storage.base_path, f"{model_id}_metadata.json")
    assert os.path.exists(pt_path)
    assert os.path.exists(json_path)
    
    # Delete model
    result = storage.delete_model(model_id)
    
    # Verify deletion was successful
    assert result is True
    
    # Verify files are removed
    assert not os.path.exists(pt_path), ".pt file should be deleted"
    assert not os.path.exists(json_path), ".json file should be deleted"


def test_delete_model_returns_false_when_no_files(storage):
    """Test that delete_model returns False when model doesn't exist."""
    result = storage.delete_model("non-existent-model")
    assert result is False


def test_delete_model_handles_partial_deletion(storage, sample_model, sample_metadata):
    """Test that delete_model handles cases where only one file exists."""
    model_id = "test-model-005"
    
    # Save checkpoint
    storage.save_checkpoint(
        model=sample_model,
        metadata=sample_metadata,
        model_id=model_id
    )
    
    # Manually delete .pt file
    pt_path = os.path.join(storage.base_path, f"{model_id}.pt")
    os.remove(pt_path)
    
    # Delete model (should still work and delete .json)
    result = storage.delete_model(model_id)
    assert result is True
    
    # Verify .json is deleted
    json_path = os.path.join(storage.base_path, f"{model_id}_metadata.json")
    assert not os.path.exists(json_path)


# ============================================================================
# TEST 5: export_to_excel creates valid Excel with all sheets
# ============================================================================

def test_export_to_excel_creates_valid_excel_with_all_sheets(storage, sample_model, sample_metadata):
    """Test that export_to_excel creates valid Excel file with all required sheets."""
    model_id = "test-model-006"
    
    # Save checkpoint first
    storage.save_checkpoint(
        model=sample_model,
        metadata=sample_metadata,
        model_id=model_id
    )
    
    # Export to Excel
    output_path = os.path.join(storage.base_path, "test_export.xlsx")
    result_path = storage.export_to_excel(model_id, output_path)
    
    # Verify file was created
    assert os.path.exists(output_path), f"Excel file not found at {output_path}"
    assert result_path == output_path
    
    # Load and verify Excel file
    wb = load_workbook(output_path)
    
    # Verify all required sheets exist
    expected_sheets = ["Summary", "Architecture", "Hyperparameters", "Training History", "Final Metrics", "Scorecard"]
    actual_sheets = wb.sheetnames
    
    for sheet_name in expected_sheets:
        assert sheet_name in actual_sheets, f"Sheet '{sheet_name}' not found in Excel file"
    
    # Verify Summary sheet has content
    ws_summary = wb["Summary"]
    assert ws_summary['A1'].value is not None, "Summary sheet should have title"
    assert "RIFT" in str(ws_summary['A1'].value).upper() or "Neural Network" in str(ws_summary['A1'].value)
    
    # Verify Architecture sheet has content
    ws_arch = wb["Architecture"]
    assert ws_arch['A1'].value is not None, "Architecture sheet should have content"
    
    # Verify Hyperparameters sheet has content
    ws_hyper = wb["Hyperparameters"]
    assert ws_hyper['A1'].value is not None, "Hyperparameters sheet should have content"
    
    # Verify Training History sheet has headers
    ws_hist = wb["Training History"]
    assert ws_hist['A1'].value is not None, "Training History sheet should have headers"
    assert "Epoch" in str(ws_hist['A1'].value) or ws_hist['A1'].value == "Epoch"
    
    # Verify Final Metrics sheet has content
    ws_metrics = wb["Final Metrics"]
    assert ws_metrics['A1'].value is not None, "Final Metrics sheet should have content"
    
    # Verify Scorecard sheet has content
    ws_score = wb["Scorecard"]
    assert ws_score['A1'].value is not None, "Scorecard sheet should have content"
    
    wb.close()


def test_export_to_excel_raises_error_for_nonexistent_model(storage):
    """Test that export_to_excel raises error for non-existent model."""
    output_path = os.path.join(storage.base_path, "test_export.xlsx")
    
    with pytest.raises(FileNotFoundError, match="Model not found"):
        storage.export_to_excel("non-existent-model", output_path)


def test_export_to_excel_contains_correct_data(storage, sample_model):
    """Test that export_to_excel contains correct data from metadata."""
    model_id = "test-model-007"
    metadata = create_mock_metadata(model_id=model_id)
    
    # Save checkpoint
    storage.save_checkpoint(
        model=sample_model,
        metadata=metadata,
        model_id=model_id
    )
    
    # Export to Excel
    output_path = os.path.join(storage.base_path, "test_export.xlsx")
    storage.export_to_excel(model_id, output_path)
    
    # Load Excel
    wb = load_workbook(output_path)
    
    # Verify Summary sheet contains model_id
    ws_summary = wb["Summary"]
    summary_values = [cell.value for row in ws_summary.iter_rows() for cell in row]
    assert model_id in [str(v) for v in summary_values if v is not None]
    
    # Verify Architecture sheet contains architecture info
    ws_arch = wb["Architecture"]
    arch_values = [str(cell.value).lower() for row in ws_arch.iter_rows() for cell in row if cell.value is not None]
    assert any("neural" in v or "network" in v or "relu" in v for v in arch_values)
    
    # Verify Training History has epoch data
    ws_hist = wb["Training History"]
    # Check if there are data rows (beyond header)
    has_data = False
    for row_idx, row in enumerate(ws_hist.iter_rows(min_row=2), start=2):
        if any(cell.value is not None for cell in row):
            has_data = True
            break
    assert has_data, "Training History should contain epoch data"
    
    wb.close()


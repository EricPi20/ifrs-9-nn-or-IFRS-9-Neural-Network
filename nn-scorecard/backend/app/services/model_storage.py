import torch
import torch.nn as nn
import json
import os
from datetime import datetime
from typing import Dict, Any, Optional, List
from dataclasses import asdict
import logging

from app.config import settings

logger = logging.getLogger(__name__)


class ModelStorage:
    """
    Save and load model checkpoints with complete metadata.
    
    Checkpoint includes:
    - Model weights
    - ALL hyperparameters
    - Training history
    - Final metrics
    - Scorecard
    - Feature information
    """
    
    def __init__(self, base_path: str = None):
        self.base_path = base_path or settings.MODEL_DIR
        os.makedirs(self.base_path, exist_ok=True)
    
    def save_checkpoint(
        self,
        model: nn.Module,
        metadata: 'CompleteModelMetadata',
        model_id: str
    ) -> str:
        """
        Save model with complete metadata.
        
        Args:
            model: Trained PyTorch model
            metadata: Complete metadata including all hyperparameters
            model_id: Unique model identifier
        
        Returns:
            Path to saved checkpoint
        """
        checkpoint_path = os.path.join(self.base_path, f"{model_id}.pt")
        
        checkpoint = {
            # Model weights
            'model_state_dict': model.state_dict(),
            
            # Complete metadata (ALL hyperparameters)
            'metadata': asdict(metadata),
            
            # Quick access fields
            'model_id': model_id,
            'segment': metadata.segment,
            'created_at': metadata.created_at,
            
            # Version info
            'rift_version': '4.0',
            'pytorch_version': torch.__version__,
        }
        
        torch.save(checkpoint, checkpoint_path)
        logger.info(f"Model saved to {checkpoint_path}")
        
        # Also save JSON metadata
        json_path = os.path.join(self.base_path, f"{model_id}_metadata.json")
        with open(json_path, 'w') as f:
            json.dump(asdict(metadata), f, indent=2, default=str)
        
        return checkpoint_path
    
    def load_checkpoint(
        self,
        model_id: str,
        model_class: type = None
    ) -> Dict[str, Any]:
        """
        Load model checkpoint.
        
        Args:
            model_id: Model identifier
            model_class: Optional model class to instantiate
        
        Returns:
            Dictionary with model and metadata
        """
        checkpoint_path = os.path.join(self.base_path, f"{model_id}.pt")
        
        if not os.path.exists(checkpoint_path):
            raise FileNotFoundError(f"Model not found: {model_id}")
        
        checkpoint = torch.load(checkpoint_path, map_location='cpu', weights_only=False)
        
        # Recreate model if class provided
        model = None
        if model_class:
            arch = checkpoint['metadata']['architecture']
            model = model_class(
                input_dim=arch['input_dim'],
                hidden_layers=arch['hidden_layers'],
                activation=arch['activation_function'],
                dropout_rate=checkpoint['metadata']['regularization']['dropout_rate'],
                use_batch_norm=arch['use_batch_normalization']
            )
            model.load_state_dict(checkpoint['model_state_dict'])
        
        return {
            'model': model,
            'model_state_dict': checkpoint['model_state_dict'],
            'metadata': checkpoint['metadata'],
            'model_id': checkpoint['model_id'],
            'segment': checkpoint['segment']
        }
    
    def list_models(self, segment: str = None) -> List[Dict]:
        """List all saved models."""
        models = []
        
        for filename in os.listdir(self.base_path):
            if filename.endswith('_metadata.json'):
                json_path = os.path.join(self.base_path, filename)
                with open(json_path) as f:
                    meta = json.load(f)
                
                if segment is None or meta.get('segment') == segment:
                    models.append({
                        'model_id': meta.get('model_id'),
                        'segment': meta.get('segment'),
                        'created_at': meta.get('created_at'),
                        'auc_roc': meta.get('final_metrics', {}).get('discrimination', {}).get('auc_roc'),
                        'gini_ar': meta.get('final_metrics', {}).get('discrimination', {}).get('gini_ar')
                    })
        
        return sorted(models, key=lambda x: x.get('created_at', ''), reverse=True)
    
    def delete_model(self, model_id: str) -> bool:
        """Delete a model checkpoint."""
        pt_path = os.path.join(self.base_path, f"{model_id}.pt")
        json_path = os.path.join(self.base_path, f"{model_id}_metadata.json")
        
        deleted = False
        for path in [pt_path, json_path]:
            if os.path.exists(path):
                os.remove(path)
                deleted = True
        
        return deleted
    
    def export_to_excel(
        self,
        model_id: str,
        output_path: str
    ) -> str:
        """
        Export model metadata to Excel for documentation.
        
        Creates sheets:
        1. Summary
        2. Architecture
        3. Hyperparameters
        4. Training History
        5. Final Metrics
        6. Scorecard
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        checkpoint = self.load_checkpoint(model_id)
        meta = checkpoint['metadata']
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Summary"
        ws['A1'] = "RIFT Neural Network Scorecard Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        summary_data = [
            ('Model ID', meta.get('model_id')),
            ('Segment', meta.get('segment')),
            ('Created', meta.get('created_at')),
            ('', ''),
            ('Test AUC', meta.get('final_metrics', {}).get('discrimination', {}).get('auc_roc')),
            ('Test Gini/AR', meta.get('final_metrics', {}).get('discrimination', {}).get('gini_ar')),
            ('Test KS', meta.get('final_metrics', {}).get('discrimination', {}).get('ks_statistic')),
        ]
        for i, (label, value) in enumerate(summary_data, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
        
        # Sheet 2: Architecture
        ws_arch = wb.create_sheet("Architecture")
        arch = meta.get('architecture', {})
        arch_data = [
            ('Model Type', arch.get('model_type')),
            ('Input Dimension', arch.get('input_dim')),
            ('Hidden Layers', str(arch.get('hidden_layers'))),
            ('Activation', arch.get('activation_function')),
            ('Batch Normalization', arch.get('use_batch_normalization')),
            ('Total Parameters', arch.get('total_parameters')),
        ]
        for i, (label, value) in enumerate(arch_data, start=1):
            ws_arch[f'A{i}'] = label
            ws_arch[f'B{i}'] = value
        
        # Sheet 3: Hyperparameters
        ws_hyper = wb.create_sheet("Hyperparameters")
        row = 1
        
        # Regularization
        ws_hyper[f'A{row}'] = "REGULARIZATION"
        ws_hyper[f'A{row}'].font = Font(bold=True)
        row += 1
        reg = meta.get('regularization', {})
        for key, value in reg.items():
            ws_hyper[f'A{row}'] = key
            ws_hyper[f'B{row}'] = value
            row += 1
        
        row += 1
        ws_hyper[f'A{row}'] = "LOSS FUNCTION"
        ws_hyper[f'A{row}'].font = Font(bold=True)
        row += 1
        loss = meta.get('loss_function', {})
        for key, value in loss.items():
            ws_hyper[f'A{row}'] = key
            ws_hyper[f'B{row}'] = str(value)
            row += 1
        
        row += 1
        ws_hyper[f'A{row}'] = "OPTIMIZER"
        ws_hyper[f'A{row}'].font = Font(bold=True)
        row += 1
        opt = meta.get('optimizer', {})
        for key, value in opt.items():
            ws_hyper[f'A{row}'] = key
            ws_hyper[f'B{row}'] = str(value)
            row += 1
        
        row += 1
        ws_hyper[f'A{row}'] = "TRAINING"
        ws_hyper[f'A{row}'].font = Font(bold=True)
        row += 1
        train = meta.get('training', {})
        for key, value in train.items():
            ws_hyper[f'A{row}'] = key
            ws_hyper[f'B{row}'] = str(value)
            row += 1
        
        # Sheet 4: Training History
        ws_hist = wb.create_sheet("Training History")
        headers = ['Epoch', 'Train Loss', 'Test Loss', 'Train AUC', 'Test AUC', 
                   'Train AR', 'Test AR', 'Train KS', 'Test KS', 'LR']
        for col, h in enumerate(headers, start=1):
            ws_hist.cell(row=1, column=col, value=h)
            ws_hist.cell(row=1, column=col).font = Font(bold=True)
        
        history = meta.get('training_history', {}).get('epochs', [])
        for i, ep in enumerate(history, start=2):
            ws_hist.cell(row=i, column=1, value=ep.get('epoch'))
            ws_hist.cell(row=i, column=2, value=ep.get('train_loss'))
            ws_hist.cell(row=i, column=3, value=ep.get('test_loss'))
            ws_hist.cell(row=i, column=4, value=ep.get('train_auc'))
            ws_hist.cell(row=i, column=5, value=ep.get('test_auc'))
            ws_hist.cell(row=i, column=6, value=ep.get('train_ar'))
            ws_hist.cell(row=i, column=7, value=ep.get('test_ar'))
            ws_hist.cell(row=i, column=8, value=ep.get('train_ks'))
            ws_hist.cell(row=i, column=9, value=ep.get('test_ks'))
            ws_hist.cell(row=i, column=10, value=ep.get('learning_rate'))
        
        # Sheet 5: Final Metrics
        ws_metrics = wb.create_sheet("Final Metrics")
        metrics = meta.get('final_metrics', {})
        row = 1
        for category, values in metrics.items():
            ws_metrics[f'A{row}'] = category.upper()
            ws_metrics[f'A{row}'].font = Font(bold=True)
            row += 1
            if isinstance(values, dict):
                for k, v in values.items():
                    ws_metrics[f'A{row}'] = k
                    ws_metrics[f'B{row}'] = v
                    row += 1
            row += 1
        
        # Sheet 6: Scorecard
        ws_score = wb.create_sheet("Scorecard")
        scorecard = meta.get('scorecard_output', {})
        ws_score['A1'] = f"Scorecard - {meta.get('segment')}"
        ws_score['A1'].font = Font(size=14, bold=True)
        ws_score['A3'] = "Base Points"
        ws_score['B3'] = scorecard.get('base_points')
        
        row = 5
        for feature in scorecard.get('features', []):
            ws_score[f'A{row}'] = feature.get('feature_name')
            ws_score[f'A{row}'].font = Font(bold=True)
            ws_score[f'B{row}'] = f"Weight: {feature.get('model_weight', 0):.4f}"
            row += 1
            
            ws_score[f'A{row}'] = "Bin"
            ws_score[f'B{row}'] = "WoE"
            ws_score[f'C{row}'] = "Points"
            row += 1
            
            for bin_info in feature.get('bins', []):
                ws_score[f'A{row}'] = bin_info.get('bin_label')
                ws_score[f'B{row}'] = bin_info.get('woe_value')
                ws_score[f'C{row}'] = bin_info.get('points')
                row += 1
            row += 1
        
        wb.save(output_path)
        return output_path

